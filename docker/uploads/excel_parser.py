import zipfile
import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl import load_workbook

def extract_checked_checkboxes(xlsx_path):
    checked_cells = []
    drawing_files_to_check = ['xl/drawings/drawing1.xml', 'xl/drawings/drawing2.xml']

    with zipfile.ZipFile(xlsx_path, 'r') as z:
        # Parcourir les fichiers drawings (parfois plusieurs)
        for drawing_file in drawing_files_to_check:
            if drawing_file in z.namelist():
                with z.open(drawing_file) as f:
                    tree = ET.parse(f)
                    root = tree.getroot()
                    ns = {
                        'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'
                    }

                    # Parcourir chaque forme ancrée
                    for anchor in root.findall('xdr:twoCellAnchor', ns):
                        cell_info = anchor.find('xdr:from', ns)
                        if cell_info is not None:
                            col = int(cell_info.find('xdr:col', ns).text)
                            row = int(cell_info.find('xdr:row', ns).text)

                            shape = anchor.find('.//a:txBody', ns)
                            if shape is not None:
                                text = ''.join(shape.itertext()).strip()
                                # Debug : afficher tout le texte trouvé
                                # print(f"Texte en {chr(ord('A') + col)}{row+1} : '{text}'")
                                if any(s in text for s in ['☒', '✓', 'Checked']):
                                    col_letter = chr(ord('A') + col)
                                    cell = f"{col_letter}{row+1}"
                                    checked_cells.append(cell)
    return checked_cells

def parse_excel_with_checkboxes(file_path, decalage_lignes, sheet_quoitester):
    checked_cells = extract_checked_checkboxes(file_path)
    print(f"Cases cochées détectées : {checked_cells}")  # Debug

    wb = load_workbook(filename=file_path, data_only=True)
    if sheet_quoitester not in wb.sheetnames:
        print(f"❌ Onglet '{sheet_quoitester}' introuvable dans {file_path}")
        return None
    sheet = wb[sheet_quoitester]

    data = list(sheet.values)
    if not data:
        return None
    columns = data[0]
    rows = data[1:]
    df = pd.DataFrame(rows, columns=columns).dropna(how='all').reset_index(drop=True)

    # Renommer la première colonne si vide
    col0 = str(df.columns[0]).strip().lower()
    if col0 in ["none", "nan", "na"]:
        df = df.rename(columns={df.columns[0]: "Famille Technique"})
    df.iloc[:, 0] = df.iloc[:, 0].ffill()

    df["Demande ICP"] = "Non"
    df["Soft de Test"] = "Non"

    for index in df.index:
        excel_row = index + 2 + decalage_lignes
        cell_f = f"F{excel_row}"
        cell_g = f"G{excel_row}"
        print(f"Index {index} -> Test des cellules {cell_f} et {cell_g}")  # Debug

        if cell_f in checked_cells:
            df.at[index, "Demande ICP"] = "Oui"
        if cell_g in checked_cells:
            df.at[index, "Soft de Test"] = "Oui"

    if "Commentaire" in df.columns:
        df["Commentaire"] = df["Commentaire"].fillna("Pas de commentaire")

    df.dropna(axis=1, how='all', inplace=True)
    return df

# Exemple d'utilisation :
if __name__ == "__main__":
    fichier_excel = "chemin/vers/ton_fichier.xlsx"
    decalage_lignes = -4  # Ajuste selon ta structure Excel
    sheet_quoitester = "Quoi tester"

    df_result = parse_excel_with_checkboxes(fichier_excel, decalage_lignes, sheet_quoitester)
    if df_result is not None:
        print(df_result)
    else:
        print("Erreur ou feuille introuvable.")
